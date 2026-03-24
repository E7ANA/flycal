"""Import data from נתונים שחף.xlsx into the scheduler database.

Grades ט-יב only. Filters:
- Skip rows with no subject, no teacher, 0 hours, שהייה, פרטני
- Single-track groupings treated as regular lessons
- Split subject names (מתמטיקה א1/א2/ב') grouped into one cluster
- Duplicate tracks deduplicated
"""

import re
from collections import defaultdict

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models import (
    ClassGroup,
    Grade,
    GroupingCluster,
    School,
    Subject,
    SubjectRequirement,
    Teacher,
    Track,
    cluster_source_classes,
    teacher_subjects,
)

EXCEL_PATH = "/Users/assulineran/Downloads/נתונים שחף.xlsx"

# Grades to import
VALID_GRADES = {"ט", "י", "יא", "יב"}
GRADE_LEVELS = {"ט": 9, "י": 10, "יא": 11, "יב": 12}

# Base subject grouping patterns — subjects that should be clustered together
# Maps a regex pattern to the base subject name for the cluster
SUBJECT_CLUSTER_PATTERNS = [
    (re.compile(r"^מתמטיקה\s"), "מתמטיקה"),
    (re.compile(r"^אנגלית\s"), "אנגלית"),
    (re.compile(r"^אנגלית\s*[אב]"), "אנגלית"),
    (re.compile(r"^תנ\"?ך\s"), 'תנ"ך'),
]


def get_base_subject(name: str) -> str | None:
    """If subject is a level-variant (e.g. 'מתמטיקה 5 יח'), return base name."""
    for pattern, base in SUBJECT_CLUSTER_PATTERNS:
        if pattern.search(name):
            return base
    return None


def get_grade_from_class(cls_name: str) -> str | None:
    for g in ["יב'", "יא'", "י'", "ט'"]:
        if g in cls_name:
            return g.rstrip("'")
    return None


def parse_excel() -> list[dict]:
    """Parse Excel and return filtered records."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["גיליון1"]

    records = []
    current_class = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]

        if vals[0] and "ריכוז שיבוץ" in str(vals[0]):
            current_class = str(vals[0]).replace("ריכוז שיבוץ לכיתה ", "").strip("'").strip()
            continue

        if vals[0] == "שעות":
            continue

        if current_class and vals[0] is not None and isinstance(vals[0], (int, float)):
            hours = int(vals[0])
            subject = vals[2]
            is_grouped = vals[3] == "כן"
            teacher = vals[4]
            lesson_type = vals[7]

            # Filters
            if not subject or hours == 0:
                continue
            if lesson_type in ("שהייה", "פרטני"):
                continue
            if not teacher:
                continue

            grade = get_grade_from_class(current_class)
            if grade not in VALID_GRADES:
                continue

            records.append(
                {
                    "class": current_class,
                    "grade": grade,
                    "hours": hours,
                    "subject": str(subject).strip(),
                    "grouped": is_grouped,
                    "teacher": str(teacher).strip(),
                }
            )

    return records


def identify_grouping_clusters(records: list[dict]) -> dict[tuple[str, str], list[dict]]:
    """Identify real grouping clusters.

    A cluster is a set of tracks (teacher+hours combos) for a subject group
    within a grade that has >1 unique teacher and runs simultaneously.

    Two types of groupings:
    1. Same subject name, multiple teachers (e.g., "מדעים" with 3 teachers in grade ט)
    2. Split subject names with a common base (e.g., "מתמטיקה א1", "מתמטיקה א2" in grade ט)
    """
    # First: group by (grade, subject_name) for same-name groupings
    same_name = defaultdict(list)
    for r in records:
        if r["grouped"]:
            same_name[(r["grade"], r["subject"])].append(r)

    # Second: group by (grade, base_subject) for split-name groupings
    split_name = defaultdict(list)
    for r in records:
        base = get_base_subject(r["subject"])
        if base and r["grouped"]:
            split_name[(r["grade"], base)].append(r)

    # Merge: if a base subject group has >1 unique subject name, it's a split grouping
    clusters = {}

    # Process split-name clusters first (they take priority)
    used_keys = set()
    for (grade, base), items in split_name.items():
        unique_subjects = set(item["subject"] for item in items)
        unique_teachers = set(item["teacher"] for item in items)
        if len(unique_subjects) > 1 and len(unique_teachers) > 1:
            # Deduplicate: unique (teacher, subject, hours) combos
            seen = set()
            deduped = []
            for item in items:
                key = (item["teacher"], item["subject"], item["hours"])
                if key not in seen:
                    seen.add(key)
                    deduped.append(item)
            clusters[(grade, base)] = deduped
            # Mark individual subject keys as used
            for subj in unique_subjects:
                used_keys.add((grade, subj))

    # Process same-name clusters (if not already handled as split)
    for (grade, subject), items in same_name.items():
        if (grade, subject) in used_keys:
            continue
        unique_teachers = set(item["teacher"] for item in items)
        if len(unique_teachers) > 1:
            # Deduplicate
            seen = set()
            deduped = []
            for item in items:
                key = (item["teacher"], item["hours"])
                if key not in seen:
                    seen.add(key)
                    deduped.append(item)
            clusters[(grade, subject)] = deduped

    return clusters


def import_data():
    records = parse_excel()
    print(f"Parsed {len(records)} records from Excel")

    clusters = identify_grouping_clusters(records)
    print(f"Identified {len(clusters)} grouping clusters")

    # Collect all cluster record keys for exclusion from regular requirements
    cluster_record_keys = set()
    for items in clusters.values():
        for item in items:
            cluster_record_keys.add((item["grade"], item["subject"], item["teacher"]))

    db: Session = SessionLocal()
    try:
        # 1. Create school
        school = School(
            name="אולפנת אלומה שדות נגב",
            days_per_week=5,
            periods_per_day=10,
            period_duration_minutes=45,
            break_slots=[],
            week_start_day="SUNDAY",
        )
        db.add(school)
        db.flush()
        school_id = school.id
        print(f"Created school: {school.name} (id={school_id})")

        # 2. Create grades
        grade_map = {}  # grade_name -> Grade obj
        for grade_name in sorted(VALID_GRADES, key=lambda g: GRADE_LEVELS[g]):
            grade = Grade(
                school_id=school_id,
                name=grade_name,
                level=GRADE_LEVELS[grade_name],
            )
            db.add(grade)
            db.flush()
            grade_map[grade_name] = grade
        print(f"Created {len(grade_map)} grades")

        # 3. Create class groups
        class_map = {}  # class_name -> ClassGroup obj
        class_names = sorted(set(r["class"] for r in records))
        for cls_name in class_names:
            grade_name = get_grade_from_class(cls_name)
            cg = ClassGroup(
                school_id=school_id,
                name=cls_name,
                grade_id=grade_map[grade_name].id,
            )
            db.add(cg)
            db.flush()
            class_map[cls_name] = cg
        print(f"Created {len(class_map)} class groups")

        # 4. Create unique subjects
        subject_names = sorted(set(r["subject"] for r in records))
        # Also add base subjects for clusters if not already present
        for (grade, base_name), items in clusters.items():
            unique_subj_names = set(item["subject"] for item in items)
            if len(unique_subj_names) > 1 and base_name not in subject_names:
                subject_names.append(base_name)
        subject_names = sorted(set(subject_names))

        subject_map = {}  # subject_name -> Subject obj
        for subj_name in subject_names:
            subj = Subject(school_id=school_id, name=subj_name)
            db.add(subj)
            db.flush()
            subject_map[subj_name] = subj
        print(f"Created {len(subject_map)} subjects")

        # 5. Create unique teachers
        teacher_names = sorted(set(r["teacher"] for r in records))
        teacher_map = {}  # teacher_name -> Teacher obj
        for t_name in teacher_names:
            teacher = Teacher(
                school_id=school_id,
                name=t_name,
                max_hours_per_week=40,
            )
            db.add(teacher)
            db.flush()
            teacher_map[t_name] = teacher
        print(f"Created {len(teacher_map)} teachers")

        # 6. Set teacher-subject qualifications
        teacher_subject_pairs = set()
        for r in records:
            teacher_subject_pairs.add((r["teacher"], r["subject"]))
        for t_name, s_name in teacher_subject_pairs:
            db.execute(
                teacher_subjects.insert().values(
                    teacher_id=teacher_map[t_name].id,
                    subject_id=subject_map[s_name].id,
                )
            )
        print(f"Created {len(teacher_subject_pairs)} teacher-subject qualifications")

        # 7. Create grouping clusters + tracks
        cluster_obj_map = {}  # (grade, base_name) -> GroupingCluster
        for (grade, cluster_name), items in clusters.items():
            unique_subj_names = set(item["subject"] for item in items)
            is_split = len(unique_subj_names) > 1

            # For split clusters, use base subject; for same-name, use the subject itself
            if is_split:
                cluster_subject = subject_map[cluster_name]
            else:
                cluster_subject = subject_map[cluster_name]

            gc = GroupingCluster(
                school_id=school_id,
                name=f"הקבצת {cluster_name} {grade}",
                subject_id=cluster_subject.id,
            )
            db.add(gc)
            db.flush()

            # Add source classes
            grade_classes = [cg for name, cg in class_map.items() if get_grade_from_class(name) == grade]
            for cg in grade_classes:
                db.execute(
                    cluster_source_classes.insert().values(
                        cluster_id=gc.id,
                        class_group_id=cg.id,
                    )
                )

            # Create tracks
            for item in items:
                track_name = item["subject"] if is_split else item["teacher"]
                track = Track(
                    name=track_name,
                    cluster_id=gc.id,
                    teacher_id=teacher_map[item["teacher"]].id,
                    hours_per_week=item["hours"],
                )
                db.add(track)

            cluster_obj_map[(grade, cluster_name)] = gc
            db.flush()

        print(f"Created {len(cluster_obj_map)} grouping clusters")

        # 8. Create regular (non-grouped) SubjectRequirements
        # For each class, add requirements for non-cluster records
        req_count = 0
        seen_reqs = set()
        for r in records:
            key = (r["grade"], r["subject"], r["teacher"])
            if key in cluster_record_keys:
                continue
            # Deduplicate same class+subject+teacher (keep first, sum hours if needed)
            req_key = (r["class"], r["subject"], r["teacher"])
            if req_key in seen_reqs:
                continue
            seen_reqs.add(req_key)

            sr = SubjectRequirement(
                school_id=school_id,
                class_group_id=class_map[r["class"]].id,
                subject_id=subject_map[r["subject"]].id,
                teacher_id=teacher_map[r["teacher"]].id,
                hours_per_week=r["hours"],
                is_grouped=False,
                grouping_cluster_id=None,
            )
            db.add(sr)
            req_count += 1

        db.flush()
        print(f"Created {req_count} regular subject requirements")

        db.commit()
        print("\nImport complete!")

        # Summary
        print(f"\n=== Summary ===")
        print(f"School: {school.name}")
        print(f"Grades: {len(grade_map)}")
        print(f"Classes: {len(class_map)}")
        print(f"Teachers: {len(teacher_map)}")
        print(f"Subjects: {len(subject_map)}")
        print(f"Regular requirements: {req_count}")
        print(f"Grouping clusters: {len(cluster_obj_map)}")
        for (grade, name), gc in sorted(cluster_obj_map.items(), key=lambda x: (GRADE_LEVELS.get(x[0][0], 0), x[0][1])):
            tracks = db.query(Track).filter(Track.cluster_id == gc.id).all()
            print(f"  {gc.name}: {len(tracks)} tracks")
            for t in tracks:
                teacher_name = teacher_map_inv.get(t.teacher_id, "?") if t.teacher_id else "?"
                print(f"    {t.name} | {t.hours_per_week} שעות | {teacher_name}")

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    # Build inverse teacher map for summary
    records = parse_excel()
    teacher_map_inv = {}
    db_temp = SessionLocal()
    try:
        import_data()
    finally:
        db_temp.close()
