"""
Import school timetable data from the Excel planning file.
Creates: School, Grades, ClassGroups, Subjects, SubjectRequirements, GroupingClusters.
"""

import sys
import os

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from sqlalchemy import text
from app.database import engine, SessionLocal, Base
from app.models.school import School, WeekStartDay
from app.models.class_group import Grade, ClassGroup, GroupingCluster, Track
from app.models.subject import Subject, SubjectRequirement
from app.models.teacher import Teacher
from app.models.constraint import Constraint
from app.models.timetable import Solution, ScheduledLesson
from app.models.timeslot import TimeSlot
from app.models.room import Room

EXCEL_PATH = "/Users/assulineran/Downloads/קובץ תיכנון 4 שנתי לפ'ה.xlsx"

# Subject colors for nice UI display
SUBJECT_COLORS = {
    "חינוך": "#EC4899",       # pink
    "אמונה": "#8B5CF6",       # purple
    "תורה": "#6366F1",        # indigo
    "נביא": "#818CF8",        # light indigo
    'תנ"ך': "#7C3AED",        # violet
    'תושב"ע': "#A78BFA",      # light violet
    "משנה": "#C084FC",        # lilac
    "רב": "#9333EA",          # deep purple
    "מנהל": "#6B7280",        # gray
    "חברותות": "#D946EF",     # fuchsia
    "עברית": "#EF4444",       # red
    "לשון": "#EF4444",        # red
    "ספרות": "#F87171",       # light red
    "הסטוריה": "#F59E0B",     # amber
    'תוע"י': "#F59E0B",       # amber
    "אזרחות": "#D97706",      # dark amber
    "מדעים": "#10B981",       # emerald
    "אנגלית": "#3B82F6",      # blue
    "מתמטיקה": "#EF4444",     # red
    "חנ": "#22C55E",          # green
    "תעבורתי": "#64748B",     # slate
    "כישורי": "#14B8A6",      # teal
    "מדעי": "#0EA5E9",        # sky
    "תאטרון": "#F472B6",      # pink
    "ארץ": "#84CC16",         # lime
    "מצויינות": "#FBBF24",    # yellow
    "קבלת שבת": "#A855F7",    # purple
    "מחול": "#FB923C",        # orange
    "מדריכה": "#94A3B8",      # gray-blue
}


def get_color(subject_name: str) -> str:
    """Match subject name to a color."""
    for key, color in SUBJECT_COLORS.items():
        if key in subject_name:
            return color
    return "#6B7280"  # default gray


def clear_all_data(db):
    """Delete all existing data from all tables."""
    print("Clearing all existing data...")
    # Order matters due to foreign keys
    db.query(ScheduledLesson).delete()
    db.query(Solution).delete()
    db.query(SubjectRequirement).delete()
    db.query(Track).delete()
    db.query(GroupingCluster).delete()
    db.query(Constraint).delete()
    db.query(TimeSlot).delete()
    db.query(Room).delete()
    # Clear teacher_subjects association
    db.execute(text("DELETE FROM teacher_subjects"))
    # Clear cluster_source_classes association
    db.execute(text("DELETE FROM cluster_source_classes"))
    db.query(ClassGroup).delete()
    db.query(Grade).delete()
    db.query(Teacher).delete()
    db.query(Subject).delete()
    db.query(School).delete()
    db.commit()
    print("  All data cleared.")


def create_school(db) -> School:
    """Create the school record."""
    school = School(
        name="אולפנה",
        days_per_week=6,
        periods_per_day=8,
        period_duration_minutes=45,
        break_slots=[],
        week_start_day=WeekStartDay.SUNDAY,
        periods_per_day_map={
            "SUNDAY": 8,
            "MONDAY": 8,
            "TUESDAY": 8,
            "WEDNESDAY": 8,
            "THURSDAY": 8,
            "FRIDAY": 5,
        },
    )
    db.add(school)
    db.flush()
    print(f"  Created school: {school.name} (id={school.id})")
    return school


def create_grades(db, school_id: int) -> dict:
    """Create grade records. Returns {name: Grade}."""
    grades_data = [
        ("ט", 9),
        ("י", 10),
        ("יא", 11),
        ("יב", 12),
    ]
    grades = {}
    for name, level in grades_data:
        grade = Grade(school_id=school_id, name=name, level=level)
        db.add(grade)
        db.flush()
        grades[name] = grade
        print(f"  Created grade: {name} (level={level}, id={grade.id})")
    return grades


def create_class_groups(db, school_id: int, grades: dict) -> dict:
    """
    Create class groups based on Excel parallels structure.
    Returns {key: ClassGroup} where key = "grade_class" e.g. "ט_1", "יא_מחא".
    """
    classes = {}

    # Grade ט: 2 parallel regular classes
    for i in range(1, 3):
        name = f"ט{i}"
        cg = ClassGroup(school_id=school_id, name=name, grade_id=grades["ט"].id, num_students=30)
        db.add(cg)
        db.flush()
        classes[f"ט_{i}"] = cg
        print(f"  Created class: {name} (id={cg.id})")

    # Grade י: 2 parallel regular classes
    for i in range(1, 3):
        name = f"י{i}"
        cg = ClassGroup(school_id=school_id, name=name, grade_id=grades["י"].id, num_students=30)
        db.add(cg)
        db.flush()
        classes[f"י_{i}"] = cg
        print(f"  Created class: {name} (id={cg.id})")

    # Grade יא: מח'א, מח'ב, ח"מ
    for key, display_name in [("מחא", "יא-מח'א"), ("מחב", "יא-מח'ב"), ("חמ", 'יא-ח"מ')]:
        cg = ClassGroup(school_id=school_id, name=display_name, grade_id=grades["יא"].id, num_students=30)
        db.add(cg)
        db.flush()
        classes[f"יא_{key}"] = cg
        print(f"  Created class: {display_name} (id={cg.id})")

    # Grade יב: מח'א, מח'ב
    for key, display_name in [("מחא", "יב-מח'א"), ("מחב", "יב-מח'ב")]:
        cg = ClassGroup(school_id=school_id, name=display_name, grade_id=grades["יב"].id, num_students=30)
        db.add(cg)
        db.flush()
        classes[f"יב_{key}"] = cg
        print(f"  Created class: {display_name} (id={cg.id})")

    return classes


def parse_excel():
    """
    Parse the Excel file and return structured data.
    Returns list of dicts with subject info and hours per class/grade.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["סרגל פדגוגי"]

    # Skip metadata rows
    SKIP_SUBJECTS = {"מקבילות", "סה\"כ", "סה\"כ לתשלום ללא תל\"ן"}

    subjects = []
    for row_num in range(4, ws.max_row + 1):
        subject_name = ws.cell(row=row_num, column=1).value
        if subject_name is None:
            continue
        subject_name = subject_name.strip()
        if subject_name in SKIP_SUBJECTS or subject_name.startswith("סה\"כ"):
            continue

        def val(col):
            v = ws.cell(row=row_num, column=col).value
            if v is None:
                return 0
            try:
                return int(v)
            except (ValueError, TypeError):
                return 0

        subject_data = {
            "name": subject_name,
            "row": row_num,
            # Grade ט - regular hours per class
            "ט_reg": val(2),      # col B
            "ט_layer": val(4),    # col D - grouping hours
            "ט_total": val(7),    # col G - total for grade
            # Grade י
            "י_reg": val(8),      # col H
            "י_layer": val(10),   # col J
            "י_total": val(13),   # col M
            # Grade יא
            "יא_מחא": val(14),    # col N
            "יא_מחב": val(15),    # col O
            "יא_חמ": val(16),     # col P
            "יא_layer": val(17),  # col Q
            "יא_total": val(20),  # col T
            # Grade יב
            "יב_מחא": val(21),    # col U
            "יב_מחב": val(22),    # col V
            "יב_חמ": val(23),     # col W (all 0 in this file)
            "יב_layer": val(24),  # col X
            "יב_total": val(27),  # col AA
        }

        # Check if this subject has any hours at all
        has_hours = any([
            subject_data["ט_reg"], subject_data["ט_layer"],
            subject_data["י_reg"], subject_data["י_layer"],
            subject_data["יא_מחא"], subject_data["יא_מחב"], subject_data["יא_חמ"], subject_data["יא_layer"],
            subject_data["יב_מחא"], subject_data["יב_מחב"], subject_data["יב_חמ"], subject_data["יב_layer"],
        ])

        if has_hours:
            subjects.append(subject_data)

    return subjects


def create_subjects(db, school_id: int, subject_data_list: list) -> dict:
    """Create Subject records. Returns {name: Subject}."""
    subjects = {}
    for sd in subject_data_list:
        name = sd["name"]
        color = get_color(name)
        subj = Subject(school_id=school_id, name=name, color=color)
        db.add(subj)
        db.flush()
        subjects[name] = subj
        print(f"  Created subject: {name} (id={subj.id}, color={color})")
    return subjects


def create_requirements(db, school_id: int, subject_data_list: list,
                        subjects: dict, classes: dict) -> int:
    """
    Create SubjectRequirement records for regular (per-class) hours.
    Returns count of created requirements.
    """
    count = 0

    for sd in subject_data_list:
        subj = subjects[sd["name"]]

        # Map: class_key -> hours from Excel
        requirements_map = {
            # Grade ט: both classes get same regular hours
            "ט_1": sd["ט_reg"],
            "ט_2": sd["ט_reg"],
            # Grade י: both classes get same regular hours
            "י_1": sd["י_reg"],
            "י_2": sd["י_reg"],
            # Grade יא: different tracks get different hours
            "יא_מחא": sd["יא_מחא"],
            "יא_מחב": sd["יא_מחב"],
            "יא_חמ": sd["יא_חמ"],
            # Grade יב: different tracks get different hours
            "יב_מחא": sd["יב_מחא"],
            "יב_מחב": sd["יב_מחב"],
        }

        for class_key, hours in requirements_map.items():
            if hours > 0 and class_key in classes:
                req = SubjectRequirement(
                    school_id=school_id,
                    class_group_id=classes[class_key].id,
                    subject_id=subj.id,
                    teacher_id=None,  # No teacher assigned yet
                    hours_per_week=hours,
                    is_grouped=False,
                    grouping_cluster_id=None,
                )
                db.add(req)
                count += 1

    db.flush()
    return count


def create_grouping_clusters(db, school_id: int, subject_data_list: list,
                              subjects: dict, classes: dict) -> int:
    """
    Create GroupingCluster records for subjects with positive layer hours.
    Creates empty clusters (no tracks yet - user will configure those).
    Returns count of created clusters.
    """
    count = 0

    # Grade mapping: which classes are the source for groupings in each grade
    grade_classes = {
        "ט": ["ט_1", "ט_2"],
        "י": ["י_1", "י_2"],
        "יא": ["יא_מחא", "יא_מחב", "יא_חמ"],
        "יב": ["יב_מחא", "יב_מחב"],
    }

    grade_layer_keys = {
        "ט": "ט_layer",
        "י": "י_layer",
        "יא": "יא_layer",
        "יב": "יב_layer",
    }

    for sd in subject_data_list:
        subj = subjects[sd["name"]]

        for grade_name, layer_key in grade_layer_keys.items():
            layer_hours = sd[layer_key]
            if layer_hours > 0:
                cluster_name = f"הקבצת {sd['name']} {grade_name}"
                cluster = GroupingCluster(
                    school_id=school_id,
                    name=cluster_name,
                    subject_id=subj.id,
                )
                db.add(cluster)
                db.flush()

                # Link source classes
                for class_key in grade_classes[grade_name]:
                    if class_key in classes:
                        cluster.source_classes.append(classes[class_key])

                # Create a single placeholder track with total hours
                track = Track(
                    name=f"{sd['name']} - הקבצה ({layer_hours} ש\"ש)",
                    cluster_id=cluster.id,
                    teacher_id=None,
                    hours_per_week=layer_hours,
                )
                db.add(track)
                db.flush()

                count += 1
                print(f"  Created grouping: {cluster_name} ({layer_hours} hours, id={cluster.id})")

    return count


def main():
    print("=" * 60)
    print("  Importing Excel data into scheduler database")
    print("=" * 60)

    # Parse Excel
    print("\n1. Parsing Excel file...")
    subject_data_list = parse_excel()
    print(f"  Found {len(subject_data_list)} subjects with hours")

    db = SessionLocal()
    try:
        # Clear existing data
        print("\n2. Clearing existing data...")
        clear_all_data(db)

        # Create school
        print("\n3. Creating school...")
        school = create_school(db)

        # Create grades
        print("\n4. Creating grades...")
        grades = create_grades(db, school.id)

        # Create class groups
        print("\n5. Creating class groups...")
        classes = create_class_groups(db, school.id, grades)

        # Create subjects
        print("\n6. Creating subjects...")
        subjects = create_subjects(db, school.id, subject_data_list)

        # Create subject requirements (regular hours)
        print("\n7. Creating subject requirements (regular hours per class)...")
        req_count = create_requirements(db, school.id, subject_data_list, subjects, classes)
        print(f"  Created {req_count} subject requirements")

        # Create grouping clusters (layer hours)
        print("\n8. Creating grouping clusters (הקבצות)...")
        cluster_count = create_grouping_clusters(db, school.id, subject_data_list, subjects, classes)
        print(f"  Created {cluster_count} grouping clusters")

        # Commit everything
        db.commit()

        # Summary
        print("\n" + "=" * 60)
        print("  IMPORT COMPLETE!")
        print("=" * 60)
        print(f"  School:              1")
        print(f"  Grades:              {len(grades)}")
        print(f"  Class groups:        {len(classes)}")
        print(f"  Subjects:            {len(subjects)}")
        print(f"  Subject requirements: {req_count}")
        print(f"  Grouping clusters:   {cluster_count}")

        # Print per-class hour summary
        print("\n  Hours per class (regular only):")
        for class_key, cg in sorted(classes.items()):
            reqs = db.query(SubjectRequirement).filter_by(class_group_id=cg.id).all()
            total = sum(r.hours_per_week for r in reqs)
            print(f"    {cg.name}: {total} hours/week ({len(reqs)} subjects)")

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
