"""Export endpoints for timetable solutions — Excel format.

Generates a professional multi-sheet workbook:
  1. סיכום  — Summary stats (classes, teachers, meetings, score)
  2. One sheet per class  — Full weekly timetable with subject colors
  3. One sheet per teacher — Lessons + meetings + track info
  4. ישיבות — All meetings: times, participants, locked teachers
"""

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models.class_group import ClassGroup, Grade, GroupingCluster, Track
from app.models.meeting import Meeting
from app.models.school import School
from app.models.subject import Subject
from app.models.teacher import Teacher
from app.models.timetable import ScheduledLesson, ScheduledMeeting, Solution

router = APIRouter(prefix="/api", tags=["export"], dependencies=[Depends(get_current_user)])

DAY_LABELS = {
    "SUNDAY": "ראשון",
    "MONDAY": "שני",
    "TUESDAY": "שלישי",
    "WEDNESDAY": "רביעי",
    "THURSDAY": "חמישי",
    "FRIDAY": "שישי",
}

DAYS_ORDER = ["SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

# ── Shared styles ────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
HEADER_FONT = Font(bold=True, size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, name="Calibri")
SUBTITLE_FONT = Font(bold=True, size=12, name="Calibri", color="444444")

# Color palette — light pastel tones
HEADER_CLASS_FILL = PatternFill("solid", fgColor="FFB4C7E7")  # Light blue
HEADER_TEACHER_FILL = PatternFill("solid", fgColor="FFC5E0B4")  # Light green
HEADER_MEETING_FILL = PatternFill("solid", fgColor="FFD5C8F0")  # Light purple
HEADER_SUMMARY_FILL = PatternFill("solid", fgColor="FFFFE5A0")  # Light gold
MEETING_CELL_FILL = PatternFill("solid", fgColor="FFF0E8FA")  # Very light purple
TRACK_CELL_FILL = PatternFill("solid", fgColor="FFFFFAEB")  # Very light yellow
EMPTY_FILL = PatternFill("solid", fgColor="FFF9F9F9")  # Very light gray
HEADER_FONT_DARK = Font(bold=True, size=11, name="Calibri", color="333333")
PERIOD_FONT = Font(bold=True, size=11, name="Calibri", color="555555")


from app.utils.colors import resolve_color_bg


def _hex_to_fill(hex_color: str) -> PatternFill | None:
    """Convert a color key or legacy hex to an openpyxl PatternFill."""
    bg = resolve_color_bg(hex_color)
    try:
        h = bg.lstrip("#")
        if len(h) == 6:
            return PatternFill("solid", fgColor=f"FF{h.upper()}")
    except Exception:
        pass
    return None


def _styled_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Write a cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment or CENTER_ALIGN
    cell.border = border or THIN_BORDER
    return cell


def _write_day_headers(ws, days, start_row, header_fill, header_font_style=None):
    """Write period + day headers for a timetable grid."""
    font = header_font_style or HEADER_FONT_DARK
    _styled_cell(ws, start_row, 1, "שעה", font=font, fill=header_fill)
    for col_idx, day in enumerate(days, start=2):
        _styled_cell(ws, start_row, col_idx, DAY_LABELS.get(day, day),
                      font=font, fill=header_fill)


# ── Summary sheet ────────────────────────────────────────────────────────

def _build_summary_sheet(
    wb: Workbook,
    solution: Solution,
    school: School,
    classes: list[ClassGroup],
    teachers: list[Teacher],
    lessons: list[ScheduledLesson],
    scheduled_meetings: list[ScheduledMeeting],
    meeting_map: dict[int, Meeting],
    grade_map: dict[int, Grade],
) -> None:
    ws = wb.create_sheet(title="סיכום", index=0)
    ws.sheet_view.rightToLeft = True

    # Title
    _styled_cell(ws, 1, 1, f"סידור מערכת — {school.name}", font=TITLE_FONT,
                  fill=None, border=None)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    row = 3
    stats = [
        ("סטטוס פתרון", solution.status),
        ("ציון כולל", f"{solution.total_score:.1f}"),
        ("זמן פתרון", f"{solution.solve_time_seconds:.1f} שניות"),
        ("כיתות", len(classes)),
        ("מורים", len(teachers)),
        ("שיעורים משובצים", len(lessons)),
        ("ישיבות משובצות", len(scheduled_meetings)),
    ]
    for label, value in stats:
        _styled_cell(ws, row, 1, label, font=HEADER_FONT, fill=HEADER_SUMMARY_FILL,
                      alignment=RIGHT_ALIGN)
        _styled_cell(ws, row, 2, str(value), font=Font(size=11, name="Calibri"),
                      alignment=RIGHT_ALIGN)
        row += 1

    # Classes list grouped by grade
    row += 1
    _styled_cell(ws, row, 1, "כיתות לפי שכבה", font=SUBTITLE_FONT, fill=None, border=None)
    row += 1

    # Group classes by grade
    grade_classes: dict[int, list[ClassGroup]] = {}
    for cg in classes:
        grade_classes.setdefault(cg.grade_id, []).append(cg)

    for grade_id, cgs in sorted(grade_classes.items(),
                                  key=lambda x: grade_map[x[0]].level
                                  if x[0] in grade_map else 0):
        grade = grade_map.get(grade_id)
        grade_name = grade.name if grade else "?"
        class_names = ", ".join(c.name for c in sorted(cgs, key=lambda c: c.name))
        _styled_cell(ws, row, 1, f"שכבה {grade_name}", font=HEADER_FONT,
                      alignment=RIGHT_ALIGN, border=None)
        _styled_cell(ws, row, 2, class_names, font=Font(size=11, name="Calibri"),
                      alignment=RIGHT_ALIGN, border=None)
        row += 1

    # Meetings summary
    row += 1
    _styled_cell(ws, row, 1, "ישיבות", font=SUBTITLE_FONT, fill=None, border=None)
    row += 1

    meeting_ids_scheduled = {sm.meeting_id for sm in scheduled_meetings}
    for mid in meeting_ids_scheduled:
        meeting = meeting_map.get(mid)
        if not meeting:
            continue
        slots = [sm for sm in scheduled_meetings if sm.meeting_id == mid]
        slot_strs = [
            f"{DAY_LABELS.get(sm.day, sm.day)} שעה {sm.period}" for sm in slots
        ]
        _styled_cell(ws, row, 1, meeting.name, font=HEADER_FONT,
                      alignment=RIGHT_ALIGN, border=None)
        _styled_cell(ws, row, 2, " | ".join(slot_strs),
                      font=Font(size=11, name="Calibri"),
                      alignment=RIGHT_ALIGN, border=None)
        _styled_cell(ws, row, 3, f"{len(meeting.teachers)} משתתפים",
                      font=Font(size=11, name="Calibri", color="666666"),
                      alignment=RIGHT_ALIGN, border=None)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ── Class sheet ──────────────────────────────────────────────────────────

def _build_class_sheet(
    wb: Workbook,
    class_group: ClassGroup,
    lessons: list[ScheduledLesson],
    subject_map: dict[int, Subject],
    teacher_map: dict[int, Teacher],
    track_map: dict[int, Track],
    days: list[str],
    max_period: int,
) -> None:
    ws = wb.create_sheet(title=class_group.name[:31])
    ws.sheet_view.rightToLeft = True

    _write_day_headers(ws, days, 1, HEADER_CLASS_FILL)

    ws.column_dimensions["A"].width = 6
    for col_idx in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Build grid — handle multiple lessons per slot (tracks)
    grid: dict[tuple[str, int], list[ScheduledLesson]] = {}
    for lesson in lessons:
        grid.setdefault((lesson.day, lesson.period), []).append(lesson)

    for period in range(0, max_period + 1):
        row = period + 2
        _styled_cell(ws, row, 1, period, font=PERIOD_FONT)

        for col_idx, day in enumerate(days, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

            slot_lessons = grid.get((day, period), [])
            if not slot_lessons:
                cell.fill = EMPTY_FILL
                continue

            # Deduplicate: if multiple lessons from same track at same slot
            # for same class, show the track name
            lines = []
            fill_color = None
            for lesson in slot_lessons:
                subj = subject_map.get(lesson.subject_id)
                teacher = teacher_map.get(lesson.teacher_id)
                subj_name = subj.name if subj else ""
                teacher_name = teacher.name if teacher else ""

                track = track_map.get(lesson.track_id) if lesson.track_id else None
                if track:
                    line = f"{subj_name} ({track.name})\n{teacher_name}"
                else:
                    line = f"{subj_name}\n{teacher_name}"
                lines.append(line)

                if not fill_color and subj and subj.color:
                    fill_color = _hex_to_fill(subj.color)

            cell.value = "\n".join(lines)
            if fill_color:
                cell.fill = fill_color


# ── Teacher sheet ────────────────────────────────────────────────────────

def _build_teacher_sheet(
    wb: Workbook,
    teacher: Teacher,
    lessons: list[ScheduledLesson],
    subject_map: dict[int, Subject],
    class_map: dict[int, ClassGroup],
    track_map: dict[int, Track],
    days: list[str],
    max_period: int,
    teacher_meetings: list[tuple[ScheduledMeeting, Meeting]] | None = None,
) -> None:
    ws = wb.create_sheet(title=f"מורה - {teacher.name}"[:31])
    ws.sheet_view.rightToLeft = True

    _write_day_headers(ws, days, 1, HEADER_TEACHER_FILL)

    ws.column_dimensions["A"].width = 6
    for col_idx in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Build lesson grid — teacher may have multiple lessons same slot (co-teaching)
    grid: dict[tuple[str, int], list[ScheduledLesson]] = {}
    for lesson in lessons:
        grid.setdefault((lesson.day, lesson.period), []).append(lesson)

    meeting_grid: dict[tuple[str, int], str] = {}
    if teacher_meetings:
        for sm, meeting in teacher_meetings:
            meeting_grid[(sm.day, sm.period)] = meeting.name

    for period in range(0, max_period + 1):
        row = period + 2
        _styled_cell(ws, row, 1, period, font=PERIOD_FONT)

        for col_idx, day in enumerate(days, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

            slot_lessons = grid.get((day, period), [])
            meeting_name = meeting_grid.get((day, period))

            if slot_lessons:
                lines = []
                for lesson in slot_lessons:
                    subj = subject_map.get(lesson.subject_id)
                    cg = class_map.get(lesson.class_group_id) if lesson.class_group_id else None
                    subj_name = subj.name if subj else ""
                    class_name = cg.name if cg else ""

                    track = track_map.get(lesson.track_id) if lesson.track_id else None
                    if track:
                        lines.append(f"{subj_name} ({track.name})\n{class_name}")
                    else:
                        lines.append(f"{subj_name}\n{class_name}")

                cell.value = "\n".join(lines)
                if slot_lessons[0].track_id:
                    cell.fill = TRACK_CELL_FILL
            elif meeting_name:
                cell.value = meeting_name
                cell.fill = MEETING_CELL_FILL
                cell.font = Font(bold=True, size=10, name="Calibri", color="7C3AED")
            else:
                cell.fill = EMPTY_FILL

    # Summary row at bottom
    row = max_period + 3
    total_lessons = len(lessons)
    unique_days = len({l.day for l in lessons})
    meetings_count = len(teacher_meetings) if teacher_meetings else 0
    _styled_cell(ws, row, 1, "סה״כ", font=HEADER_FONT, fill=HEADER_TEACHER_FILL,
                  alignment=RIGHT_ALIGN)
    ws.merge_cells(start_row=row, start_column=2, end_row=row,
                   end_column=len(days) + 1)
    summary = f"{total_lessons} שיעורים | {unique_days} ימי הוראה"
    if meetings_count:
        summary += f" | {meetings_count} ישיבות"
    _styled_cell(ws, row, 2, summary, font=Font(size=11, name="Calibri"),
                  fill=None, alignment=RIGHT_ALIGN, border=None)


# ── Meetings sheet ───────────────────────────────────────────────────────

def _build_meetings_sheet(
    wb: Workbook,
    scheduled_meetings: list[ScheduledMeeting],
    meeting_map: dict[int, Meeting],
    teacher_map: dict[int, Teacher],
    days: list[str],
    max_period: int,
    all_lessons: list[ScheduledLesson] | None = None,
) -> None:
    ws = wb.create_sheet(title="ישיבות")
    ws.sheet_view.rightToLeft = True

    # Gather scheduled meeting IDs
    meeting_ids = sorted({sm.meeting_id for sm in scheduled_meetings})
    if not meeting_ids:
        return

    # ── Part 1: Meeting timetable grid ──
    _styled_cell(ws, 1, 1, "לוח ישיבות", font=TITLE_FONT, fill=None, border=None)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 1)

    _write_day_headers(ws, days, 2, HEADER_MEETING_FILL)
    ws.column_dimensions["A"].width = 6

    # Build grid: (day, period) -> list of meeting names
    mtg_grid: dict[tuple[str, int], list[str]] = {}
    for sm in scheduled_meetings:
        meeting = meeting_map.get(sm.meeting_id)
        name = meeting.name if meeting else f"#{sm.meeting_id}"
        mtg_grid.setdefault((sm.day, sm.period), []).append(name)

    for period in range(0, max_period + 1):
        row = period + 2
        _styled_cell(ws, row, 1, period, font=PERIOD_FONT)
        for col_idx, day in enumerate(days, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            names = mtg_grid.get((day, period), [])
            if names:
                cell.value = "\n".join(names)
                cell.fill = MEETING_CELL_FILL
                cell.font = Font(bold=True, size=10, name="Calibri", color="7C3AED")
            else:
                cell.fill = EMPTY_FILL

    for col_idx in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # ── Part 2: Per-meeting participant list ──
    base_row = max_period + 4
    for mid in meeting_ids:
        meeting = meeting_map.get(mid)
        if not meeting:
            continue

        _styled_cell(ws, base_row, 1, meeting.name, font=SUBTITLE_FONT,
                      fill=HEADER_MEETING_FILL, alignment=RIGHT_ALIGN)

        # Slot info
        slots = [sm for sm in scheduled_meetings if sm.meeting_id == mid]
        slot_str = ", ".join(
            f"{DAY_LABELS.get(sm.day, sm.day)} שעה {sm.period}" for sm in slots
        )
        _styled_cell(ws, base_row, 2, slot_str,
                      font=Font(size=11, name="Calibri"),
                      fill=HEADER_MEETING_FILL, alignment=RIGHT_ALIGN)

        from app.models.meeting import MeetingType
        is_plenary = meeting.meeting_type == MeetingType.PLENARY.value
        type_labels = {
            "HOMEROOM": "מחנכות", "COORDINATORS": "רכזים",
            "MANAGEMENT": "ניהול", "CUSTOM": "מותאם אישית", "PLENARY": "מליאה",
        }
        type_label = type_labels.get(meeting.meeting_type, meeting.meeting_type)
        mandatory_label = "חובה" if meeting.is_mandatory_attendance else "לא חובה"
        _styled_cell(ws, base_row, 3, f"{type_label} | {mandatory_label}",
                      font=Font(size=10, name="Calibri", color="666666"),
                      fill=HEADER_MEETING_FILL, alignment=RIGHT_ALIGN)

        ws.merge_cells(start_row=base_row, start_column=3,
                       end_row=base_row, end_column=len(days) + 1)
        base_row += 1

        # For plenary: determine which preferred teachers attend (teach on plenary day)
        plenary_days: set[str] = set()
        teacher_teach_days: dict[int, set[str]] = {}
        if is_plenary and all_lessons:
            for sm in scheduled_meetings:
                if sm.meeting_id == mid:
                    plenary_days.add(sm.day)
            for lesson in all_lessons:
                teacher_teach_days.setdefault(lesson.teacher_id, set()).add(lesson.day)

        # Table headers
        locked_ids = set(meeting.locked_teacher_ids or [])
        _styled_cell(ws, base_row, 1, "#", font=HEADER_FONT, fill=EMPTY_FILL)
        _styled_cell(ws, base_row, 2, "שם מורה", font=HEADER_FONT, fill=EMPTY_FILL)
        _styled_cell(ws, base_row, 3, "סטטוס", font=HEADER_FONT, fill=EMPTY_FILL)
        if is_plenary:
            _styled_cell(ws, base_row, 4, "נוכחות", font=HEADER_FONT, fill=EMPTY_FILL)
        base_row += 1

        for idx, teacher in enumerate(meeting.teachers, start=1):
            t = teacher_map.get(teacher.id, teacher)
            is_locked = teacher.id in locked_ids

            if is_plenary:
                if is_locked:
                    status = "נוכחות חובה"
                    status_color = "DC2626"
                else:
                    status = "נוכחות מועדפת"
                    status_color = "2563EB"

                # Attendance: does teacher teach on plenary day?
                t_days = teacher_teach_days.get(teacher.id, set())
                if is_locked:
                    attendance = "נוכח/ת"
                    att_color = "16A34A"
                elif plenary_days & t_days:
                    attendance = "נוכח/ת"
                    att_color = "16A34A"
                else:
                    attendance = "לא נוכח/ת"
                    att_color = "EA580C"
            else:
                status = "נעול — חייב להשתתף" if is_locked else "משתתף"
                status_color = "B45309" if is_locked else "666666"
                attendance = None
                att_color = None

            font_style = Font(bold=True, size=11, name="Calibri") if is_locked else Font(size=11, name="Calibri")
            _styled_cell(ws, base_row, 1, idx, font=Font(size=11, name="Calibri"))
            _styled_cell(ws, base_row, 2, t.name, font=font_style, alignment=RIGHT_ALIGN)
            _styled_cell(ws, base_row, 3, status, font=Font(size=10, name="Calibri",
                          color=status_color), alignment=RIGHT_ALIGN)
            if attendance is not None:
                _styled_cell(ws, base_row, 4, attendance,
                              font=Font(bold=True, size=10, name="Calibri", color=att_color),
                              alignment=RIGHT_ALIGN)
            base_row += 1

        base_row += 1  # Gap between meetings

    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16


# ── Main export endpoint ─────────────────────────────────────────────────

@router.get("/solutions/{solution_id}/export/excel")
def export_excel(solution_id: int, db: Session = Depends(get_db)):
    """Export a solution as a professional Excel workbook."""
    solution = db.get(Solution, solution_id)
    if not solution:
        raise HTTPException(status_code=404, detail="פתרון לא נמצא")

    school = db.get(School, solution.school_id)
    if not school:
        raise HTTPException(status_code=404, detail="בית ספר לא נמצא")

    # Load all data
    lessons = (
        db.query(ScheduledLesson)
        .filter(ScheduledLesson.solution_id == solution_id)
        .all()
    )
    scheduled_meetings = (
        db.query(ScheduledMeeting)
        .filter(ScheduledMeeting.solution_id == solution_id)
        .all()
    )
    all_meetings = db.query(Meeting).filter(Meeting.school_id == school.id).all()
    meeting_map = {m.id: m for m in all_meetings}

    subjects = db.query(Subject).filter(Subject.school_id == school.id).all()
    subject_map = {s.id: s for s in subjects}

    teachers = db.query(Teacher).filter(Teacher.school_id == school.id).all()
    teacher_map = {t.id: t for t in teachers}

    classes = db.query(ClassGroup).filter(ClassGroup.school_id == school.id).all()
    class_map = {c.id: c for c in classes}

    grades = db.query(Grade).filter(Grade.school_id == school.id).all()
    grade_map = {g.id: g for g in grades}

    clusters = db.query(GroupingCluster).filter(GroupingCluster.school_id == school.id).all()
    track_map: dict[int, Track] = {}
    for cluster in clusters:
        for track in cluster.tracks:
            track_map[track.id] = track

    # Build teacher -> meetings lookup
    teacher_meeting_lookup: dict[int, list[tuple[ScheduledMeeting, Meeting]]] = {}
    for sm in scheduled_meetings:
        meeting = meeting_map.get(sm.meeting_id)
        if not meeting:
            continue
        for t in meeting.teachers:
            teacher_meeting_lookup.setdefault(t.id, []).append((sm, meeting))

    days = DAYS_ORDER[: school.days_per_week]
    max_period = school.periods_per_day

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Summary sheet
    _build_summary_sheet(
        wb, solution, school, classes, teachers, lessons,
        scheduled_meetings, meeting_map, grade_map,
    )

    # 2. Per-class sheets (sorted by grade level, then name)
    sorted_classes = sorted(
        classes,
        key=lambda c: (grade_map[c.grade_id].level
                       if c.grade_id in grade_map else 0, c.name),
    )
    for cg in sorted_classes:
        class_lessons = [l for l in lessons if l.class_group_id == cg.id]
        if class_lessons:
            _build_class_sheet(
                wb, cg, class_lessons, subject_map, teacher_map,
                track_map, days, max_period,
            )

    # 3. Per-teacher sheets (sorted by name)
    sorted_teachers = sorted(teachers, key=lambda t: t.name)
    for teacher in sorted_teachers:
        teacher_lessons = [l for l in lessons if l.teacher_id == teacher.id]
        t_meetings = teacher_meeting_lookup.get(teacher.id)
        if teacher_lessons or t_meetings:
            _build_teacher_sheet(
                wb, teacher, teacher_lessons, subject_map, class_map,
                track_map, days, max_period, teacher_meetings=t_meetings,
            )

    # 4. Meetings sheet
    if scheduled_meetings:
        _build_meetings_sheet(
            wb, scheduled_meetings, meeting_map, teacher_map, days, max_period,
            all_lessons=lessons,
        )

    # Write to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"timetable_solution_{solution_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
