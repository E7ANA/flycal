"""Import endpoints — bulk data loading from Excel/CSV files."""

from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.class_group import ClassGroup, Grade
from app.models.subject import Subject, SubjectRequirement
from app.models.teacher import Teacher

router = APIRouter(prefix="/api/import", tags=["import"])


@router.post("/teachers")
async def import_teachers(
    school_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import teachers from Excel/CSV.

    Expected columns: שם, מקסימום_שעות, מינימום_שעות, אחוז_משרה, מקצועות
    The מקצועות column should contain comma-separated subject names.
    """
    content = await file.read()
    rows, errors = _parse_excel_or_csv(content, file.filename or "")

    if not rows:
        raise HTTPException(status_code=400, detail="הקובץ ריק או בפורמט לא נתמך")

    # Load existing subjects for matching by name
    subjects = db.query(Subject).filter(Subject.school_id == school_id).all()
    subject_by_name = {s.name.strip(): s for s in subjects}

    created = 0
    import_errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        name = _get_cell(row, ["שם", "name", "שם_מורה"])
        if not name:
            import_errors.append(f"שורה {i}: חסר שם")
            continue

        max_hours = _get_int(row, ["מקסימום_שעות", "max_hours", "שעות_מקסימום"], 40)
        min_hours = _get_int(row, ["מינימום_שעות", "min_hours", "שעות_מינימום"], None)
        employment = _get_float(row, ["אחוז_משרה", "employment", "משרה"], None)

        teacher = Teacher(
            school_id=school_id,
            name=name.strip(),
            max_hours_per_week=max_hours,
            min_hours_per_week=min_hours,
            employment_percentage=employment,
        )

        # Parse subject names
        subject_str = _get_cell(row, ["מקצועות", "subjects"])
        if subject_str:
            for sname in subject_str.split(","):
                sname = sname.strip()
                if sname in subject_by_name:
                    teacher.subjects.append(subject_by_name[sname])

        db.add(teacher)
        created += 1

    db.commit()

    return {
        "message": f"יובאו {created} מורים",
        "created": created,
        "errors": import_errors,
    }


@router.post("/classes")
async def import_classes(
    school_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import classes from Excel/CSV.

    Expected columns: שם, שכבה, מספר_תלמידים
    """
    content = await file.read()
    rows, errors = _parse_excel_or_csv(content, file.filename or "")

    if not rows:
        raise HTTPException(status_code=400, detail="הקובץ ריק או בפורמט לא נתמך")

    # Load existing grades
    grades = db.query(Grade).filter(Grade.school_id == school_id).all()
    grade_by_name = {g.name.strip(): g for g in grades}

    created = 0
    import_errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        name = _get_cell(row, ["שם", "name", "כיתה"])
        if not name:
            import_errors.append(f"שורה {i}: חסר שם")
            continue

        grade_name = _get_cell(row, ["שכבה", "grade"])
        if not grade_name or grade_name.strip() not in grade_by_name:
            import_errors.append(f"שורה {i}: שכבה '{grade_name}' לא נמצאה")
            continue

        grade = grade_by_name[grade_name.strip()]
        num_students = _get_int(row, ["מספר_תלמידים", "students", "תלמידים"], 30)

        cg = ClassGroup(
            school_id=school_id,
            name=name.strip(),
            grade_id=grade.id,
            num_students=num_students,
        )
        db.add(cg)
        created += 1

    db.commit()

    return {
        "message": f"יובאו {created} כיתות",
        "created": created,
        "errors": import_errors,
    }


@router.post("/requirements")
async def import_requirements(
    school_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import subject requirements from Excel/CSV.

    Expected columns: כיתה, מקצוע, מורה, שעות
    """
    content = await file.read()
    rows, errors = _parse_excel_or_csv(content, file.filename or "")

    if not rows:
        raise HTTPException(status_code=400, detail="הקובץ ריק או בפורמט לא נתמך")

    classes = db.query(ClassGroup).filter(ClassGroup.school_id == school_id).all()
    class_by_name = {c.name.strip(): c for c in classes}

    subjects = db.query(Subject).filter(Subject.school_id == school_id).all()
    subject_by_name = {s.name.strip(): s for s in subjects}

    teachers = db.query(Teacher).filter(Teacher.school_id == school_id).all()
    teacher_by_name = {t.name.strip(): t for t in teachers}

    created = 0
    import_errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        class_name = _get_cell(row, ["כיתה", "class"])
        subject_name = _get_cell(row, ["מקצוע", "subject"])
        teacher_name = _get_cell(row, ["מורה", "teacher"])
        hours = _get_int(row, ["שעות", "hours", "שעות_שבועיות"], None)

        if not class_name or class_name.strip() not in class_by_name:
            import_errors.append(f"שורה {i}: כיתה '{class_name}' לא נמצאה")
            continue
        if not subject_name or subject_name.strip() not in subject_by_name:
            import_errors.append(f"שורה {i}: מקצוע '{subject_name}' לא נמצא")
            continue
        if hours is None or hours <= 0:
            import_errors.append(f"שורה {i}: חסר מספר שעות")
            continue

        cg = class_by_name[class_name.strip()]
        subj = subject_by_name[subject_name.strip()]
        teacher = teacher_by_name.get(teacher_name.strip()) if teacher_name else None

        req = SubjectRequirement(
            school_id=school_id,
            class_group_id=cg.id,
            subject_id=subj.id,
            teacher_id=teacher.id if teacher else None,
            hours_per_week=hours,
            is_grouped=False,
        )
        db.add(req)
        created += 1

    db.commit()

    return {
        "message": f"יובאו {created} דרישות",
        "created": created,
        "errors": import_errors,
    }


# ─── Helpers ──────────────────────────────────────────────

def _parse_excel_or_csv(
    content: bytes, filename: str,
) -> tuple[list[dict[str, str]], list[str]]:
    """Parse an Excel or CSV file into a list of row dicts."""
    errors: list[str] = []

    if filename.endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            if ws is None:
                return [], ["הקובץ לא מכיל גליון"]
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else "" for h in next(rows_iter, [])]
            rows = []
            for r in rows_iter:
                row_dict = {}
                for h, v in zip(headers, r):
                    if h and v is not None:
                        row_dict[h] = str(v).strip()
                if any(row_dict.values()):
                    rows.append(row_dict)
            return rows, errors
        except Exception as e:
            return [], [f"שגיאה בקריאת Excel: {e}"]

    elif filename.endswith(".csv"):
        import csv
        from io import StringIO

        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            rows = [
                {k.strip(): v.strip() for k, v in row.items() if k}
                for row in reader
            ]
            return rows, errors
        except Exception as e:
            return [], [f"שגיאה בקריאת CSV: {e}"]

    return [], ["פורמט קובץ לא נתמך — יש להשתמש ב-.xlsx או .csv"]


def _get_cell(row: dict[str, str], keys: list[str]) -> str | None:
    for k in keys:
        if k in row and row[k]:
            return row[k]
    return None


def _get_int(row: dict[str, str], keys: list[str], default: int | None) -> int | None:
    val = _get_cell(row, keys)
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _get_float(row: dict[str, str], keys: list[str], default: float | None) -> float | None:
    val = _get_cell(row, keys)
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default
